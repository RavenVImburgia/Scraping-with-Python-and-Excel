#https://bit.ly/2lVhlLX
# via: https://analytics.usa.gov/
import requests
import time
import os
from openpyxl import Workbook, load_workbook

excel_file = "LiveUpdates.xlsx"

while True:
    url = 'https://analytics.usa.gov/data/live/realtime.json'
    j = requests.get(url).json()
    active_users = j['data'][0]['activeUsers']
    current_time = time.strftime("%Y-%m-%d %H:%M:%S")

    if os.path.exists(excel_file):
        wb = load_workbook(excel_file)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Live Updates"

    ws['A1'] = "Number of people visiting a U.S. government website-"
    ws['A2'] = "Active Users Right Now:"
    ws['A3'] = active_users
    ws['A4'] = f"Last updated: {current_time}"

    wb.save(excel_file)
    print("Excel file updated!")

    print("Number of people visiting a U.S. government website-")
    print("Active Users Right Now:")
    print(active_users)
    print("Updated at:", current_time)
    print("Waiting 5 minutes...\n")

    time.sleep(300)